import os
import io
import re
import boto3
import httpx
import asyncio
import openpyxl
import multiprocessing
import pandas as pd
import polars as pl
from functools import partial
from datetime import datetime, date
from app.scripts.data_fetchers.broker_data import BrokerData
from dotenv import load_dotenv
from app.logger import logger
from typing import Dict, Tuple


load_dotenv()
pd.set_option('future.no_silent_downcasting', True)

class ApiError(Exception):
    """Custom exception for API errors."""
    pass

class KeynoteApi:

    def __init__(self):
        """Initialize the API client with base URL, headers, and API key."""
        self.BASE_URL = "https://backoffice.wizzer.in/shrdbms/dotnet/api/stansoft/"
        self.HEADERS = {"Content-Type": "application/json"}
        self.API_KEY = os.getenv("SHAREPRO_WIZZER_API_KEY")

    async def fetch_holding(self, for_date: str, ucc: str) -> Dict:
        """
        Fetch holding data for a specific date and UCC.
        Returns a pandas DataFrame if successful, None otherwise.
        """
        url = self.BASE_URL + "GetDpHoldingData"
        payload = {
            "key": self.API_KEY,
            "ucc": ucc,
            "segments": "NSDL,CDSL,NFO,NSE,BSE",
            "date": datetime.strptime(for_date, "%Y-%m-%d").strftime("%d-%m-%Y")
        }
        try:
            data = await self._fetch_api_data(url, payload)
            return data.get("curdata", {})
        except ApiError as e:
            logger.error(f"Error fetching holding data: {e}")
            return None

    async def fetch_ledger(self, from_date: str, to_date: str, ucc: str) -> Dict:
        """
        Fetch ledger data between two dates for a UCC.
        Returns a pandas DataFrame if successful, None otherwise.
        """
        params = self._generate_ledger_params(from_date, to_date, ucc)
        all_transactions = []
        for p in params:
            url = self.BASE_URL + "ClientLedgerData"
            payload = {
                "key": self.API_KEY,
                "datefrom": p["datefrom"],
                "dateto": p["dateto"],
                "segments": "NSE,BSE,NFO",
                "ucc": p["ucc"],
                "accyear": p["accyear"]
            }
            try:
                data = await self._fetch_api_data(url, payload)
                if "curdata" in data:
                    all_transactions.extend(data["curdata"])
            except ApiError as e:
                logger.info(f"Error fetching ledger data for {p['datefrom']} to {p['dateto']}: {e}")
                continue
        if not all_transactions:
            return None
        
        all_transactions_df = pd.DataFrame(all_transactions)
        all_transactions_df["vrdt"] = pd.to_datetime(all_transactions_df["vrdt"]).dt.date
        return all_transactions_df.to_dict()
    
    def _generate_ledger_params(self, from_date: str, to_date: str, ucc: str):
        """
        Generate parameters for ledger API calls, splitting by financial years.
        Returns a list of parameter dictionaries.
        """
        from_date_dt = datetime.strptime(from_date, "%Y-%m-%d")
        to_date_dt = datetime.strptime(to_date, "%Y-%m-%d")
        params = []
        
        first_period_end = datetime(from_date_dt.year, 3, 31)
        if from_date_dt > first_period_end:
            first_period_end = datetime(from_date_dt.year + 1, 3, 31)
        if first_period_end > to_date_dt:
            first_period_end = to_date_dt
        
        params.append({
            "datefrom": from_date_dt.strftime("%d-%m-%Y"),
            "dateto": first_period_end.strftime("%d-%m-%Y"),
            "accyear": f"{from_date_dt.year % 100}{(from_date_dt.year + 1) % 100}",
            "ucc": ucc
        })
        
        while first_period_end < to_date_dt:
            next_period_start = datetime(first_period_end.year, 4, 1)
            next_period_end = datetime(first_period_end.year + 1, 3, 31)
            if next_period_end > to_date_dt:
                next_period_end = to_date_dt
            params.append({
                "datefrom": next_period_start.strftime("%d-%m-%Y"),
                "dateto": next_period_end.strftime("%d-%m-%Y"),
                "accyear": f"{next_period_start.year % 100}{(next_period_start.year + 1) % 100}",
                "ucc": ucc
            })
            first_period_end = next_period_end
        
        return params

    async def _fetch_api_data(self, url: str, payload: dict):
        """
        Helper function to fetch API data with retries and error handling.
        Returns a dictionary if successful, raises ApiError otherwise.
        """
        tries = 3
        for attempt in range(tries):
            try:
                async with httpx.AsyncClient(timeout=30) as client:
                    response = await client.post(url, headers=self.HEADERS, json=payload)
                    
                    if response.status_code != 200:
                        raise ApiError(f"Server error—status code {response.status_code}")
                    
                    content_type = response.headers.get('Content-Type', '').lower()
                    if 'application/json' not in content_type:
                        raise ApiError(f"Expected JSON, got {content_type or 'no content type'}")
                    
                    try:
                        data = response.json()
                    except ValueError:
                        raise ApiError("Response is not JSON")
                    
                    if not isinstance(data, dict):
                        raise ApiError(f"Expected dict, got {type(data).__name__}")
                    
                    return data
            
            except httpx.ReadTimeout:
                logger.info(f"Timeout on attempt {attempt + 1}/{tries}")
                if attempt < tries - 1:
                    await asyncio.sleep(5)
                else:
                    raise ApiError("Max retries reached, giving up")
            except httpx.HTTPStatusError as e:
                raise ApiError(f"HTTP error: {e.response.status_code} - {e.response.text}")
            except Exception as e:
                raise ApiError(f"Unexpected error: {str(e)}")


class KeynoteDataProcessor:

    def __init__(self):
        """Initialize the Keynote data processor."""
        self.bulk_ledger_location = "/home/admin/Plus91Backoffice/Plus91_Backend/data/bulk_ledgers"
        self.bulk_holdings_location = "/home/admin/Plus91Backoffice/Plus91_Backend/data/bulk_holdings"
        self.broker = BrokerData()
        self.master_data = self.broker.get_master_data()
        self.master_df = pd.DataFrame(self.master_data['data'])
        self.logger = logger

    def fetch_ledger(self, ucc: str, from_date: str = None, to_date: str = None) -> dict:
        """
        For a given client code and optional date range, scans all bulk ledger Excel files in folder_path
        and returns a consolidated Polars DataFrame (as a dictionary list) of transactions for the client.

        Date strings must be in "yyyy-mm-dd" format. If only one date is provided then:
          - If only from_date is provided, transactions from that date until the latest are returned.
          - If only to_date is provided, transactions from the earliest date until to_date are returned.
          - If both are omitted, the complete ledger is returned.
        """
        start_date = self._parse_optional_date(from_date, date.min)
        end_date   = self._parse_optional_date(to_date, date.max)

        files = [os.path.join(self.bulk_ledger_location, f) for f in os.listdir(self.bulk_ledger_location) if f.endswith(".xlsx")]
        if not files:
            raise FileNotFoundError(f"No Excel files found in folder {self.bulk_ledger_location}.")

        pool = multiprocessing.Pool()
        func = partial(self._process_file, client_code=ucc, start_date=start_date, end_date=end_date)
        results = pool.map(func, files)
        pool.close()
        pool.join()

        combined_rows = []
        header_included = False
        for res in results:
            if res:
                if not header_included:
                    combined_rows.extend(res)
                    header_included = True
                else:
                    combined_rows.extend(res[1:])  # Skip header in subsequent results

        if not combined_rows:
            self.logger.info(f"No ledger entries found for client code {ucc} in the provided date range.")
            return {}

        header = combined_rows[0]
        data_rows = combined_rows[1:]
        df = pl.DataFrame(data_rows, schema=header)
        return df.to_dicts()

    def _process_file(self, file_path, client_code, start_date, end_date):
        """
        Process a single bulk ledger file. Searches for client ledger blocks matching the given client_code
        and returns a list of rows (as lists) that fall within the date range.
        
        Expects:
          - A header row in the second row.
          - Ledger entries starting from the third row.
        """
        result_rows = []
        try:
            wb = openpyxl.load_workbook(file_path, data_only=True)
        except Exception as e:
            self.logger.error(f"Could not load {file_path}: {e}")
            return result_rows

        sheet = wb.active
        data = list(sheet.iter_rows(values_only=True))
        if len(data) < 2:
            return result_rows

        header = list(data[1])
        in_target_client = False
        ledger_block = [header]
        for row in data[2:]:
            if row[0] is not None:
                if self._is_client_header(row):
                    if in_target_client:
                        filtered = self._filter_block_by_date(ledger_block, start_date, end_date)
                        if filtered and len(filtered) > 1:
                            result_rows.extend(filtered)
                        in_target_client = False
                        ledger_block = [header]
                    if re.search(r'\[' + re.escape(client_code) + r'\]', str(row[0])):
                        in_target_client = True
                        ledger_block = [header]
                        ledger_block.append(list(row))
                    else:
                        in_target_client = False
                        ledger_block = [header]
                    continue

            if in_target_client:
                ledger_block.append(list(row))
        if in_target_client and ledger_block:
            filtered = self._filter_block_by_date(ledger_block, start_date, end_date)
            if filtered and len(filtered) > 1:
                result_rows.extend(filtered)
        return result_rows

    def _filter_block_by_date(self, ledger_block, start_date, end_date):
        """
        Given a ledger block (a list of rows with the first row as the header),
        groups rows into "transaction blocks" (a main row with a valid date plus any detail rows)
        and returns only those blocks where the main transaction date falls between start_date and end_date.
        """
        filtered = []
        header = ledger_block[0]
        filtered.append(header)

        current_block = []
        transaction_date = None

        for row in ledger_block[1:]:
            dt = self._parse_date_from_cell(row[0])
            if dt is not None:
                if current_block:
                    if transaction_date and start_date <= transaction_date <= end_date:
                        filtered.extend(current_block)
                transaction_date = dt
                current_block = [row]
            else:
                if current_block:
                    current_block.append(row)
        if current_block and transaction_date and start_date <= transaction_date <= end_date:
            filtered.extend(current_block)
        return filtered

    def _parse_date_from_cell(self, cell_value):
        """
        Attempts to parse a cell value as a date in the "dd-MMM-yyyy" format.
        Returns a datetime.date object if successful; otherwise, None.
        """
        if cell_value is None:
            return None
        try:
            if isinstance(cell_value, date):
                return cell_value
            return datetime.strptime(str(cell_value).strip(), "%d-%b-%Y").date()
        except Exception:
            return None

    def _is_client_header(self, row):
        """
        Determines if a row is a client header. A valid client header should have a non-empty
        first cell that contains a code enclosed in square brackets (e.g. "Client Name [MN025]")
        while all other cells are empty or whitespace.
        """
        if row[0] is None:
            return False
        text = str(row[0])
        if '[' in text and ']' in text:
            for cell in row[1:]:
                if cell is not None and str(cell).strip() != "":
                    return False
            return True
        return False

    def _parse_optional_date(self, date_str: str, default: date) -> date:
        """
        If date_str is provided, parses it using _validate_date_str.
        If not provided, returns the default date.
        """
        if date_str:
            return self._validate_date_str(date_str)
        return default

    def _validate_date_str(self, date_str: str) -> date:
        """
        Validates and parses a date string in the "yyyy-mm-dd" format.
        Raises ValueError if the date is invalid.
        """
        try:
            dt = datetime.strptime(date_str, "%Y-%m-%d").date()
            return dt
        except ValueError as e:
            raise ValueError(f"Invalid date string '{date_str}': {e}")

    def _extract_file_date(self, file_path: str) -> str:
        """
        Extracts the date from a bulk holdings file name.
        Expected format: "Bulk Holdings_{dd-mm-yy}.xlsx" and converts dd-mm-yy to yyyy-mm-dd.
        """
        base_name = os.path.basename(file_path)
        date_match = re.search(r'_(\d{2})-(\d{2})-(\d{2})', base_name)
        if date_match:
            day, month, year_short = date_match.groups()
            full_year = "20" + year_short  # Assumes 21st century.
            return f"{full_year}-{month}-{day}"
        else:
            raise ValueError("Could not extract a date from the file name.")

    def _extract_client_blocks(self, file_path: str) -> dict:
        """
        Reads the bulk holdings Excel file and groups rows into client blocks.
        Each block (a list of rows with the header as the first row) is keyed by the client code.
        """
        wb = openpyxl.load_workbook(file_path, data_only=True)
        sheet = wb.active
        all_rows = list(sheet.iter_rows(values_only=True))

        header = None
        header_index = None
        for i, row in enumerate(all_rows):
            if any(cell is not None and str(cell).strip() != "" for cell in row):
                header = list(row)
                header_index = i
                break
        if header is None:
            raise ValueError("No header row found in the bulk holdings file.")

        client_blocks = {}
        current_client_code = None
        current_block = None

        for row in all_rows[header_index+1:]:
            if row[0] is not None and ('[' in str(row[0]) and ']' in str(row[0])):
                if current_client_code and current_block:
                    client_blocks[current_client_code] = current_block
                match = re.search(r'\[(.*?)\]', str(row[0]))
                if match:
                    current_client_code = match.group(1).strip()
                else:
                    current_client_code = None
                current_block = [header]
                current_block.append(list(row))
            else:
                if current_block is not None:
                    current_block.append(list(row))
        if current_client_code and current_block:
            client_blocks[current_client_code] = current_block

        return client_blocks

    def _process_and_upload_client_block(self, args):
        """
        Processes a single client block from a bulk holdings file: selects and renames columns,
        cleans the trading symbol and uploads the result (as an Excel file) to S3.

        The S3 key follows the structure:
          PLUS91_PMS/ledgers_and_holdings/keynote/single_accounts/{client_code}/holdings/{file_date}.xlsx
        """
        client_code, block, file_date_str, bucket_name = args

        desired_cols = ["Scrip", "ISIN", "Margin_x000D_\nQuantity", "Market_x000D_\nValue"]
        rename_mapping = {
            "Scrip": "trading_symbol",
            "ISIN": "isin",
            "Margin_x000D_\nQuantity": "quantity",
            "Market_x000D_\nValue": "market_value"
        }

        try:
            df = pd.DataFrame(block[1:], columns=block[0])
        except Exception as ex:
            self.logger.error(f"Error creating DataFrame for client {client_code}: {ex}")
            return

        df.columns = [str(col).strip() for col in df.columns]
        missing = [col for col in desired_cols if col not in df.columns]
        if missing:
            self.logger.warning(f"For client {client_code} the following columns are missing: {missing}. Skipping.")
            return

        df = df[desired_cols].copy()
        df.rename(columns=rename_mapping, inplace=True)

        def clean_symbol(symbol):
            if pd.isna(symbol):
                return symbol
            cleaned = re.sub(r'\s+(MF|EQ|NO)$', '', str(symbol)).strip()
            return cleaned

        df["trading_symbol"] = df["trading_symbol"].apply(clean_symbol)
        df = df[df["isin"].notna()]
        df = df.groupby('trading_symbol', as_index=False).agg({
            'isin': 'first',
            'quantity': 'sum',
            'market_value': 'sum'
        })
        
        master_df = self.master_df[["isin", "trading_symbol"]]
        df = df.merge(master_df, on="isin", how="left")
        df['trading_symbol_y'] = df['trading_symbol_y'].fillna(df['isin'])

        df = df.drop(columns=["trading_symbol_x"], errors="ignore")
        df = df.rename(columns={"trading_symbol_y": "trading_symbol"})
        df = df.drop_duplicates()

        output_buffer = io.BytesIO()

        if df.empty:
            self.logger.warning(f"Empty DataFrame for client {client_code}. Uploading empty file with headers.")
            empty_df = pd.DataFrame(columns=df.columns)
            empty_df.to_excel(output_buffer, index=False)
        else:
            df.to_excel(output_buffer, index=False)

        output_buffer.seek(0)

        s3_key = f"PLUS91_PMS/ledgers_and_holdings/keynote/single_accounts/{client_code}/holdings/{file_date_str}.xlsx"
        try:
            s3_client = boto3.client('s3')
            s3_client.upload_fileobj(output_buffer, bucket_name, s3_key)
            self.logger.info(f"Uploaded file for client {client_code} to s3://{bucket_name}/{s3_key}")
        except Exception as e:
            self.logger.error(f"Error uploading file for client {client_code} to S3: {e}")
        output_buffer.close()

    def process_bulk_holdings_to_s3(self, file_path: str, bucket_name: str = "plus91backoffice"):
        """
        Processes a single bulk holdings Excel file by:
          1. Extracting the file date from its name (converted to yyyy-mm-dd).
          2. Splitting the file into client blocks.
          3. Processing each client block (selecting, renaming, cleaning columns)
             and uploading each as an Excel file to S3 in the proper folder structure.
             
        Uses multiprocessing to process client blocks concurrently.
        """
        file_date_str = self._extract_file_date(file_path)
        self.logger.info(f"Extracted file date: {file_date_str} from {file_path}")

        client_blocks = self._extract_client_blocks(file_path)
        if not client_blocks:
            self.logger.info("No client blocks found in the bulk holdings file.")
            return

        args_list = []
        for client_code, block in client_blocks.items():
            args_list.append((client_code, block, file_date_str, bucket_name))

        with multiprocessing.Pool() as pool:
            pool.map(self._process_and_upload_client_block, args_list)

        self.logger.info("Processing of bulk holdings file complete.")

    def process_all_bulk_holdings_to_s3(self, bucket_name: str = "plus91backoffice"):
        """
        Scans the given folder for bulk holdings Excel files that match the naming format:
          Bulk Holdings_{dd-mm-yy}.xlsx
        For each matching file, processes it such that individual client holding files are uploaded to S3
        under the folder structure:
          PLUS91_PMS/ledgers_and_holdings/keynote/single_accounts/{client_code}/holdings/{yyyy-mm-dd}.xlsx
        """
        file_pattern = r'^Bulk Holdings_\d{2}-\d{2}-\d{2}\.xlsx$'
        files = [f for f in os.listdir(self.bulk_holdings_location) if re.match(file_pattern, f)]
        if not files:
            self.logger.info(f"No bulk holdings files found in folder {self.bulk_holdings_location} matching the required naming format.")
            return
        
        for filename in files:
            full_path = os.path.join(self.bulk_holdings_location, filename)
            self.logger.info(f"Processing bulk holdings file: {full_path}")
            try:
                self.process_bulk_holdings_to_s3(full_path, bucket_name)
            except Exception as e:
                self.logger.error(f"Error processing file {full_path}: {e}")


class ZerodhaDataFetcher:
    def __init__(self, bucket_name="plus91backoffice", 
                 base_prefix="PLUS91_PMS/ledgers_and_holdings/zerodha/single_accounts/"):
        """Set up the S3 connection with your bucket and base prefix."""
        self.s3 = boto3.client('s3')
        self.bucket_name = bucket_name
        self.base_prefix = base_prefix.rstrip('/') + '/'

    def _get_s3_file_bytes(self, key):
        """Fetch an XLSX file from S3 and return its raw bytes."""
        full_key = self.base_prefix + key
        try:
            obj = self.s3.get_object(Bucket=self.bucket_name, Key=full_key)
            return obj['Body'].read()
        except Exception as e:
            logger.warning(f"Couldn’t grab {full_key}: {e}")
            return None

    def get_ledger(
            self, 
            broker_code: str, 
            as_dataframe=True
            ) -> Dict:
        """
        Fetch the ledger XLSX file for a broker_code from S3.
        Returns raw bytes if as_dataframe=False, or a pandas DataFrame if True.
        """
        key = f"{broker_code}/ledger/ledger-{broker_code}.xlsx"
        file_bytes = self._get_s3_file_bytes(key)
        if file_bytes is None:
            return None
        
        if not as_dataframe:
            return file_bytes
        
        df = pd.read_excel(io.BytesIO(file_bytes), header=None)
        new_column_names = {
                0: 'Index',
                1: 'Particulars',
                2: 'Posting Date',
                3: 'Cost Center',
                4: 'Voucher Type',
                5: 'Debit',
                6: 'Credit',
                7: 'Net Balance'
            }
        df = df.rename(columns=new_column_names)
        df = df.fillna('blank')
        df = df[df["Net Balance"] != "blank"].reset_index(drop=True).drop(columns=["Index"], index=0).reset_index(drop=True)
        df["Posting Date"] = pd.to_datetime(df["Posting Date"].replace("blank", pd.NaT), errors="coerce")
        df["Credit"] = pd.to_numeric(df["Credit"].replace("blank", 0), errors='coerce')
        df["Debit"] = pd.to_numeric(df["Debit"].replace("blank", 0), errors='coerce')
        df["Net Balance"] = pd.to_numeric(df["Net Balance"].replace("blank", 0), errors="coerce")
        return df.to_dict()

    def get_holdings(
            self,
            year: int,
            month: int,
            broker_code: str,
            as_dataframe=True
            ) -> Tuple[Dict, str]:
        """
        Fetch the latest holdings XLSX file for a given month and year for a broker_code.
        Returns raw bytes if as_dataframe=False, or a list of pandas DataFrames if True.
        """
        target_year_month = f"{year}-{month:02d}"
        prefix = f"{broker_code}/holdings/"
        
        response = self.s3.list_objects_v2(Bucket=self.bucket_name, 
                                          Prefix=self.base_prefix + prefix)
        if 'Contents' not in response:
            logger.warning(f"No holdings files found for {broker_code} under {self.base_prefix + prefix}!")
            return None

        holdings_files = []
        for obj in response['Contents']:
            try:
                file_date_str = obj['Key'].split('/')[-1].replace('.xlsx', '')
                file_date = datetime.strptime(file_date_str, "%Y-%m-%d")
                if file_date.strftime("%Y-%m") == target_year_month:
                    holdings_files.append((obj['Key'].replace(self.base_prefix, ''), file_date))
            except ValueError:
                logger.info(f"Skipping odd file name: {obj['Key']}")

        if not holdings_files:
            logger.warning(f"No holdings files for {target_year_month} under {broker_code}!")
            return None

        latest_file = max(holdings_files, key=lambda x: x[1])

        snapshot_date = ""
        match = re.search(r'(\d{4}-\d{2}-\d{2})', latest_file[0])
        if match:
            snapshot_date = match.group(1)
        else:
            logger.warning(f"Could not extract date from file name: {latest_file}")
            snapshot_date = None

        file_bytes = self._get_s3_file_bytes(latest_file[0])
        if file_bytes is None:
            return None
        
        if not as_dataframe:
            return file_bytes

        df = pd.read_excel(io.BytesIO(file_bytes), header=None)
        new_column_names = {
            1: 'Symbol',
            2: 'ISIN',
            3: 'Sector',
            4: 'Quantity Available',
            5: 'Quantity Discrepant',
            6: 'Quantity Long Term',
            7: 'Quantity Pledged (Margin)',
            8: 'Quantity Pledged (Loan)',
            9: 'Average Price',
            10: 'Previous Closing',
            11: 'Unrealized P&L',
            12: 'Unrealized P&L Pct'
        }
        df = df.rename(columns=new_column_names).drop(columns=0).fillna("blank")
        df = df[df["Unrealized P&L Pct"] != "blank"].reset_index(drop=True).drop(index=0).reset_index(drop=True)
        return df.to_dict(), snapshot_date


# if __name__ == "__main__":
#     fetcher = ZerodhaDataFetcher()
#     keynote = KeynoteApi()


#     # holdingsexample = asyncio.run(keynote.fetch_holding(for_date="2023-01-31", ucc="SG102"))
#     # df = pd.DataFrame(holdingsexample)
#     # print(df)

#     ledgerexample = asyncio.run(keynote.fetch_ledger(from_date="2022-04-01", to_date="2025-03-28", ucc="MK100"))
#     ledger_df = pd.DataFrame(ledgerexample)
#     ledger_df.to_csv("LEDGER_MK100_1.csv")
